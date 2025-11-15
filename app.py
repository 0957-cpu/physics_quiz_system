from flask import Flask, render_template, request, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime, date  # ✅ 一次匯入 datetime 和 date
import random
import os
import json
import gspread



from google.oauth2.service_account import Credentials

from openpyxl import load_workbook

def load_question_bank():
    """從 questions.xlsx 載入題庫，並檢查欄位完整性"""
    filename = "questions.xlsx"
    required_headers = ["id", "text", "options", "answer", "explanation", "category"]

    try:
        wb = load_workbook(filename)
        ws = wb["Questions"]
    except FileNotFoundError:
        print(f"❌ 找不到題庫檔案：{filename}")
        return []
    except KeyError:
        print("❌ 找不到工作表『Questions』，請確認 Excel 的工作表名稱。")
        return []
    except Exception as e:
        print(f"❌ 題庫載入失敗：{e}")
        return []

    # 檢查表頭欄位
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    missing_headers = [h for h in required_headers if h not in headers]

    if missing_headers:
        print(f"⚠️ 題庫缺少欄位：{', '.join(missing_headers)}")
        print(f"目前讀到的表頭：{headers}")
        return []

    # 把欄位名稱對應到欄索引
    col_idx = {h: headers.index(h) for h in required_headers}
    questions = []
    error_list = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        qid = str(row[col_idx["id"]]).strip() if row[col_idx["id"]] else ""
        text = str(row[col_idx["text"]]).strip() if row[col_idx["text"]] else ""
        options_str = row[col_idx["options"]]
        answer = str(row[col_idx["answer"]]).strip() if row[col_idx["answer"]] else ""
        explanation = str(row[col_idx["explanation"]]).strip() if row[col_idx["explanation"]] else ""
        category = str(row[col_idx["category"]]).strip() if row[col_idx["category"]] else ""

        # 檢查基本欄位是否齊全
        if not qid or not text:
            error_list.append(f"第 {i} 列：缺少題號或題目文字。")
            continue

        # 處理選項
        options = []
        if options_str:
            options = [opt.strip() for opt in str(options_str).split(",") if opt.strip()]
        if not options:
            error_list.append(f"第 {i} 列（{qid}）：沒有選項。")

        # 檢查答案是否在選項中
        if answer and options and answer not in options:
            error_list.append(f"第 {i} 列（{qid}）：答案「{answer}」不在選項中。")

        questions.append({
            "id": qid,
            "text": text,
            "options": options,
            "answer": answer,
            "explanation": explanation,
            "category": category,
        })

    # 印出載入結果與錯誤統計
    print(f"✅ 題庫載入完成，共 {len(questions)} 題。")
    if error_list:
        print("⚠️ 以下題目內容有問題：")
        for err in error_list:
            print("   -", err)
    else:
        print("🟢 題庫檢查通過，無錯誤。")

    return questions


# 啟動時載入題庫
QUESTION_BANK = load_question_bank()
SETTINGS_FILE = "settings.json"

DEFAULT_SETTINGS = {
    "questions_per_test": 5,        # 每次抽題數
    "show_explanation": True,       # 顯示詳解
    "wrong_only_mode": False,       # 錯題再練
    "daily_limit": 3,               # 每日作答次數上限（0 = 不限制）
    "time_limit_seconds": 0         # 作答時間（秒），0 表示不啟用倒數計時
}


def load_settings():
    if not os.path.exists(SETTINGS_FILE):
        save_settings(DEFAULT_SETTINGS)
        return DEFAULT_SETTINGS

    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
    except:
        save_settings(DEFAULT_SETTINGS)
        return DEFAULT_SETTINGS

    # 若有新欄位，用預設值補
    for k, v in DEFAULT_SETTINGS.items():
        if k not in data:
            data[k] = v
    return data

def save_settings(settings: dict):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

SETTINGS = load_settings()


app = Flask(__name__)
app.secret_key = "change-this-secret-key"  # 可以改成你自己的亂碼字串

USERS_FILE = "users.xlsx"
RESULT_FILE = "quiz_results.xlsx"
# ===== Google Sheets 設定 =====
import os
from google.oauth2.service_account import Credentials
import gspread

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

GOOGLE_CREDS_FILE = os.path.join(BASE_DIR, "service_account.json")  # 確保用絕對路徑
GOOGLE_SHEET_NAME = "quiz_results_online"

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

_sheet = None  # 暫存 worksheet 物件

def get_google_sheet():
    """取得 Google Sheet 的 sheet1 物件。"""
    global _sheet
    if _sheet is None:
        try:
            print("📡 正在連線到 Google 試算表…")

            creds = Credentials.from_service_account_file(
                GOOGLE_CREDS_FILE,
                scopes=GOOGLE_SCOPES
            )
            client = gspread.authorize(creds)

            # ✅ 用名稱開啟試算表
            sh = client.open(GOOGLE_SHEET_NAME)
            _sheet = sh.sheet1  # 第一個工作表

            # 🔍 除錯資訊：印出實際寫入的試算表網址與工作表名稱
            print("✅ 已連線到 Google 試算表：", sh.url)
            print("✅ 使用的工作表名稱：", _sheet.title)

        except Exception as e:
            print("❌ 連線 Google 試算表失敗：", e)
            # 這裡 raise 讓你在測試時看到錯誤（正式上線也可以改成 pass）
            raise

    return _sheet


# ===== 題庫設定 =====
# 之後你只要一直在這裡加題目就好
QUESTION_BANK = load_question_bank()


NUM_QUESTIONS_PER_QUIZ = 3  # 每次測驗抽幾題


def load_wrong_questions(account):#老師介面錯題讀取
    """從 quiz_results.xlsx 擷取該學生所有錯題 ID"""
    if not os.path.exists("quiz_results.xlsx"):
        return []

    wb = load_workbook("quiz_results.xlsx")
    ws = wb.active

    wrong_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        t, acc, name, score, qid, ans, ok = row
        if acc == account and ok == "錯誤":
            wrong_ids.add(qid)

    return [q for q in QUESTION_BANK if q["id"] in wrong_ids]


# ===== Excel 初始化 =====

def init_users_excel():
    """如果沒有 users.xlsx，就建立一份，順便放幾個測試帳號。"""
    if not os.path.exists(USERS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["帳號", "密碼", "姓名", "總積分"])
        # 測試資料：之後你可以改成真正學生名單
        ws.append(["s001", "1234", "小明", 0])
        ws.append(["s002", "1234", "小美", 0])
        wb.save(USERS_FILE)


def init_results_excel():
    """如果沒有 quiz_results.xlsx，就建立一份（每列一人一次作答）。"""
    if not os.path.exists(RESULT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"

        # 基本欄位
        headers = ["時間", "帳號", "姓名", "作答次數", "本次分數"]

        # 依照題庫動態加欄位：每題兩欄（答案 / 是否正確）
        for q in QUESTION_BANK:
            headers.append(f"{q['id']}_答案")
            headers.append(f"{q['id']}_是否正確")

        ws.append(headers)
        wb.save(RESULT_FILE)



# ===== 輔助函式 =====

def get_user_row(account):
    """回傳 (wb, ws, row_index) 讓你可以更新該使用者總積分。"""
    wb = load_workbook(USERS_FILE)
    ws = wb["Users"]
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value == account:
            return wb, ws, idx
    return wb, ws, None
def get_user_rank(account):
    """根據總積分計算該帳號的排名（1 是最高分）。"""
    wb = load_workbook(USERS_FILE)
    ws = wb["Users"]

    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc, pwd, name, total = row
        total = total or 0
        users.append((acc, total))

    # 按總積分由高到低排序
    users.sort(key=lambda x: x[1], reverse=True)

    rank = None
    for idx, (acc, _) in enumerate(users, start=1):
        if acc == account:
            rank = idx
            break

    return rank, len(users)
def get_level(total_points):
    """根據總積分回傳等級稱號。你可以自己改門檻和名稱。"""
    if total_points < 10:
        return "Lv.1再多嘗試幾次"
    elif total_points < 30:
        return "Lv.2基礎觀念get"
    elif total_points < 60:
        return "Lv.3 觀念越來越完善了"
    else:
        return "Lv.4 安心考試去"


# ===== 路由設定 =====

@app.route("/")
def index():
    if "user_account" in session:
        return redirect(url_for("home"))
    return redirect(url_for("login"))



@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        account = request.form.get("account", "")
        password = request.form.get("password", "")

        from openpyxl import load_workbook

        # 讀 users.xlsx
        wb = load_workbook(USERS_FILE, read_only=True)
        try:
            ws = wb["Users"]
        except KeyError:
            ws = wb.active  # 萬一你的表不叫 Users

        # 讀表頭並建立欄位索引（1 列）
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        need = ["account", "password", "name", "total_points"]
        miss = [h for h in need if h not in headers]
        if miss:
            return render_template("login.html", error=f"users.xlsx 缺少欄位：{', '.join(miss)}")

        col = {h: headers.index(h) for h in need}  # 0-based index

        # 掃資料列
        found = False
        user_name = ""
        total_points = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 用表頭索引安全取值（即使有多餘欄也不怕）
            acc  = str(row[col["account"]]  or "").strip()
            pwd  = str(row[col["password"]] or "").strip()
            name = str(row[col["name"]]     or "").strip()
            total = row[col["total_points"]]
            total = int(total) if isinstance(total, (int, float)) else 0

            if acc == str(account).strip() and pwd == str(password).strip():
                found = True
                user_name = name or acc
                total_points = total
                break

        if found:
            session["user_account"] = account
            session["user_name"] = user_name
            session["total_points"] = total_points
            session["logged_in"] = True
            
            # ✅ 這行很重要：標記是不是老師
            session["is_teacher"] = (account == "t001")  # ← 你的老師帳號

            if session["is_teacher"]:
                return redirect(url_for("teacher_home"))
            else:
                return redirect(url_for("home"))
            
        else:
            return render_template("login.html", error="帳號或密碼錯誤")

    return render_template("login.html")



@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.pop("user_account", None)
    session.pop("user_name", None)
    session.pop("total_points", None)
    return redirect(url_for("login"))

@app.route("/home")
def home():
    if "user_account" not in session:
        return redirect(url_for("login"))

    account = session["user_account"]
    name = session.get("user_name", "同學")
    total_points = session.get("total_points", 0)

    # 等級（你原本的等級函式）
    level = get_level(total_points)

    # 排名（你原本就有 get_user_rank）
    try:
        rank, total_users = get_user_rank(account)
    except Exception as e:
        print("計算排名時發生錯誤：", e)
        rank, total_users = None, None

    # === 今日作答上限狀態 ===
    daily_limit = SETTINGS.get("daily_limit", 0)
    today = date.today().isoformat()

    # 如果是新的一天，重置今天計次
    if session.get("last_quiz_date") != today:
        session["last_quiz_date"] = today
        session["quiz_times_today"] = 0

    used_times = session.get("quiz_times_today", 0)

    if daily_limit == 0:
        limit_msg = "今日作答不限次數。"
        reached_limit = False
        remaining = None
    else:
        if used_times >= daily_limit:
            limit_msg = f"⚠️ 您今日已達作答上限（{daily_limit} 次）。"
            reached_limit = True
            remaining = 0
        else:
            remaining = daily_limit - used_times
            limit_msg = f"今日剩餘可作答次數：{remaining} 次（上限 {daily_limit} 次）"
            reached_limit = False

    # === 從成績檔抓統計資料 ===
    today_attempts = []   # 今日作答紀錄
    total_attempts = 0    # 總作答次數
    best_score = None     # 最高分
    avg_score = None      # 平均分
    last_score = None     # 最近一次分數
    last_time = None      # 最近一次時間

    try:
        wb_r = load_workbook(RESULT_FILE)
        ws_r = wb_r["Results"]

        scores_sum = 0

        for row in ws_r.iter_rows(min_row=2, values_only=True):
            # 結構：[時間, 帳號, 姓名, 作答次數, 分數, 之後是各題答案/對錯...]
            tstr, acc, nm, attempt_no, score = row[:5]

            if acc != account:
                continue

            total_attempts += 1
            score = score or 0
            scores_sum += score

            if best_score is None or score > best_score:
                best_score = score

            # 最近一次作答（覆蓋到最後一筆）
            last_score = score
            last_time = str(tstr)

            # 今日作答紀錄
            if tstr and str(tstr).startswith(today):
                today_attempts.append({
                    "time": str(tstr),
                    "attempt_no": attempt_no,
                    "score": score
                })

        if total_attempts > 0:
            avg_score = round(scores_sum / total_attempts, 1)

        # 依時間排序今日作答紀錄
        today_attempts = sorted(today_attempts, key=lambda x: x["time"])

    except FileNotFoundError:
        # 還沒有成績檔，代表沒人作答過
        pass
    except Exception as e:
        print("讀取成績檔錯誤：", e)

    return render_template(
        "home.html",
        name=name,
        total_points=total_points,
        level=level,
        rank=rank,
        total_users=total_users,
        daily_limit=daily_limit,
        limit_msg=limit_msg,
        reached_limit=reached_limit,
        remaining=remaining,
        today_attempts=today_attempts,
        total_attempts=total_attempts,
        best_score=best_score,
        avg_score=avg_score,
        last_score=last_score,
        last_time=last_time
    )

@app.route("/teacher_home")
def teacher_home():
    # 只有老師可以看（如果老師帳號不是 t001，請改這裡）
    if session.get("user_account") != "t001" and not session.get("is_teacher"):
        return redirect(url_for("home"))

    # 讀 users.xlsx 來做排行榜
    try:
        wb = load_workbook(USERS_FILE)
        # 你原本的使用者表名稱如果是 "Users" 就用這個，若是 active 就改成 wb.active
        try:
            ws = wb["Users"]
        except KeyError:
            ws = wb.active

        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 預期欄位：[帳號, 密碼, 姓名, 總積分, ...]
            if not row or not row[0]:
                continue
            account, pwd, name, total_points = (row + (0, 0, 0, 0))[:4]
            total_points = total_points or 0

            students.append({
                "account": account,
                "name": name,
                "total_points": total_points
            })

        # 依總積分排序（大到小），若積分相同以姓名排序
        students.sort(key=lambda s: (-s["total_points"], s["name"] or ""))

        # 幫每個學生加上名次（1,2,3,...）
        for idx, s in enumerate(students, start=1):
            s["rank"] = idx

        total_students = len(students)
        avg_points = None
        max_points = None

        if total_students > 0:
            max_points = max(s["total_points"] for s in students)
            avg_points = round(
                sum(s["total_points"] for s in students) / total_students, 1
            )

    except FileNotFoundError:
        students = []
        total_students = 0
        avg_points = None
        max_points = None

    return render_template(
        "teacher_home.html",
        teacher_name=session.get("user_name", "老師"),
        students=students,
        total_students=total_students,
        avg_points=avg_points,
        max_points=max_points
    )



@app.route("/quiz")
def quiz():
    if "user_account" not in session:
        return redirect(url_for("login"))

    account = session["user_account"]

    # daily limit（教師設定）
    limit = SETTINGS.get("daily_limit", 0)
    if limit > 0:
        today = date.today().isoformat()

        # 偵測該學生今天是否已作答
        if session.get("last_quiz_date") == today:
            if session.get("quiz_times_today", 0) >= limit:
                return f"⚠️ 您今天的作答次數已達上限（{limit} 次）。"

        # 若尚未作答，初始化計次
        if session.get("last_quiz_date") != today:
            session["last_quiz_date"] = today
            session["quiz_times_today"] = 0

    # 錯題模式（教師設定）
    if SETTINGS.get("wrong_only_mode", False):
        wrong_q = load_wrong_questions(account)
        if wrong_q:
            usable_bank = wrong_q
        else:
            usable_bank = QUESTION_BANK
    else:
        usable_bank = QUESTION_BANK

    if not usable_bank:
        return "⚠️ 沒有可用的題目。"

    # 取得抽題數
    n = min(SETTINGS.get("questions_per_test", 5), len(usable_bank))
    questions_for_view = random.sample(usable_bank, n)

    # 打亂選項
    for q in questions_for_view:
        if "options" in q:
            random.shuffle(q["options"])

    return render_template(
        "quiz.html",
        name=session["user_name"],
        quiz=questions_for_view,
        show_explanation=SETTINGS.get("show_explanation", True),
        time_limit_seconds=SETTINGS.get("time_limit_seconds", 0)
)
    


@app.route("/admin")
def admin():
    """簡單老師後台：列出所有學生統計（一列一個測驗）。"""
    # 讀取 users.xlsx
    wb_u = load_workbook(USERS_FILE)
    ws_u = wb_u["Users"]

    users = []
    for row in ws_u.iter_rows(min_row=2, values_only=True):
        acc, pwd, name, total = row
        total = total or 0
        users.append({
            "account": acc,
            "name": name,
            "total_points": total,
        })

    # 讀取 quiz_results.xlsx，計算作答次數與平均分數
    wb_r = load_workbook(RESULT_FILE)
    ws_r = wb_r["Results"]

    # 準備一個 map 來累積每個人的測驗次數和總分
    stats_map = {}
    for u in users:
        stats_map[u["account"]] = {"attempts": 0, "sum_score": 0}

    for row in ws_r.iter_rows(min_row=2, values_only=True):
        # 依照欄位順序：時間, 帳號, 姓名, 作答次數, 本次分數, ...
        time_str, acc, name, attempt_no, score, *rest = row
        if acc in stats_map:
            stats_map[acc]["attempts"] += 1
            stats_map[acc]["sum_score"] += (score or 0)

    # 合併回 users
    for u in users:
        acc = u["account"]
        att = stats_map[acc]["attempts"]
        ssum = stats_map[acc]["sum_score"]
        u["attempts"] = att
        u["avg_score"] = round(ssum / att, 2) if att > 0 else None

    # 依照總積分由高到低排序
    users.sort(key=lambda x: x["total_points"], reverse=True)

    return render_template("admin.html", users=users)



@app.route("/points")
def points():
    # 沒登入就回登入頁
    if "user_account" not in session:
        return redirect(url_for("login"))
    session["logged_in"] = True

    account = session["user_account"]
    name = session["user_name"]

     # 讀取作答成績檔 quiz_results.xlsx
    wb = load_workbook(RESULT_FILE)
    ws = wb["Results"]

    # 找出該學生所有紀錄
    records = []
    # 根據你現在的一列一人格式：
    # 欄位：0時間, 1帳號, 2姓名, 3作答次數, 4本次分數, ...
    total_points = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == account:  # 帳號欄
            total_points += row[4]  # 累積分數欄
            records.append({
                "time": row[0],
                "score": row[4],
                "points": total_points,
                "rank": "-"
            })
    
    # 目前總積分 & 排名
    # 你之前有 get_user_rank / get_user_row，就直接用那個
    rank, total_users = get_user_rank(account)
    total_points = session.get("total_points", 0)

    return render_template(
        "points.html",
        name=name,
        records=records,
        total_points=total_points,
        rank=rank,
        total_users=total_users,
        title="積分查詢"
    )

@app.route("/settings", methods=["GET", "POST"])  # 老師設定
def settings_page():
    # 這裡用你的老師帳號判斷（如果不是 t001 請改成你自己的）
    if session.get("user_account") != "t001":
        return redirect(url_for("quiz"))

    global SETTINGS
    message = None
    error = None

    if request.method == "POST":
        # 1. 每次抽題數
        q_str = request.form.get("questions_per_test", "").strip()
        # 2. 顯示詳解
        show_explanation = "show_explanation" in request.form
        # 3. 錯題再練
        wrong_only_mode = "wrong_only_mode" in request.form
        # 4. 每日次數上限
        limit_str = request.form.get("daily_limit", "").strip()
        # 5. 作答時間（分鐘）
        time_limit_str = request.form.get("time_limit_minutes", "").strip()

        try:
            # 抽題數
            if not q_str:
                raise ValueError("請輸入每次測驗抽出的題數。")
            q_num = int(q_str)
            if q_num <= 0:
                raise ValueError("題數必須是大於 0 的整數。")

            # 每日上限
            if limit_str == "":
                daily_limit = 0
            else:
                daily_limit = int(limit_str)
                if daily_limit < 0:
                    raise ValueError("每日作答上限不可為負數。")

            # 作答時間（分鐘 → 秒）
            if time_limit_str == "":
                time_limit_seconds = 0
            else:
                time_limit_minutes = int(time_limit_str)
                if time_limit_minutes < 0:
                    raise ValueError("作答時間不可為負數。")
                time_limit_seconds = time_limit_minutes * 60

            # ✅ 寫回設定
            SETTINGS["questions_per_test"] = q_num
            SETTINGS["show_explanation"] = show_explanation
            SETTINGS["wrong_only_mode"] = wrong_only_mode
            SETTINGS["daily_limit"] = daily_limit
            SETTINGS["time_limit_seconds"] = time_limit_seconds

            save_settings(SETTINGS)
            message = "設定已更新 ✔"

            print("🛠 設定更新：", SETTINGS)

        except ValueError as e:
            error = str(e)

    return render_template(
        "settings.html",
        settings=SETTINGS,
        name=session.get("user_name", "老師"),
        message=message,
        error=error,
        title="老師設定"
    )



@app.route("/change_password", methods=["GET", "POST"])
def change_password():
    """學生自行變更密碼，同步更新 Excel 與 Google 試算表"""
    if "user_account" not in session:
        return redirect(url_for("login"))

    account = session["user_account"]
    name = session.get("user_name", account)
    message = None
    error = None

    if request.method == "POST":
        current = request.form.get("current_password", "")
        new1 = request.form.get("new_password", "")
        new2 = request.form.get("confirm_password", "")

        if not current or not new1 or not new2:
            error = "請完整輸入目前密碼與新密碼。"
        elif new1 != new2:
            error = "兩次輸入的新密碼不一致。"
        elif len(new1) < 4:
            error = "新密碼至少需 4 個字元。"
        else:
            from openpyxl import load_workbook
            try:
                wb_u = load_workbook(USERS_FILE)
                ws_u = wb_u["Users"]
            except Exception as e:
                return render_template("change_password.html", name=name, error=f"讀取使用者資料失敗：{e}")

            updated = False
            for row in ws_u.iter_rows(min_row=2):
                acc_cell, pwd_cell, name_cell, total_cell = row
                if str(acc_cell.value) == account:
                    if str(pwd_cell.value) != current:
                        error = "目前密碼不正確。"
                    else:
                        pwd_cell.value = new1
                        updated = True
                    break

            if updated and not error:
                try:
                    wb_u.save(USERS_FILE)
                    message = "密碼已更新成功！下次登入請使用新密碼。"

                    # === 同步更新到 Google 試算表 ===
                    try:
                        sheet = get_google_sheet()  # 你原本用來連接的函式
                        records = sheet.get_all_records()  # 取全部資料列
                        # 找到該帳號對應的列
                        row_index = None
                        for i, rec in enumerate(records, start=2):  # 第1列是表頭
                            if str(rec.get("account")) == account:
                                row_index = i
                                break
                        if row_index:
                            # 密碼欄是第2欄 (B)，若你的表格欄位不同請改這裡
                            sheet.update_cell(row_index, 2, new1)
                        else:
                            print("⚠️ Google Sheet 找不到該帳號，未更新密碼。")
                    except Exception as e:
                        print("Google Sheet 更新密碼失敗：", e)

                except PermissionError:
                    error = "無法寫入 users.xlsx（可能正在被 Excel 開啟）。請先關閉再試一次。"
                except Exception as e:
                    error = f"儲存失敗：{e}"

    return render_template("change_password.html", name=name, message=message, error=error, title="變更密碼")

@app.route("/submit", methods=["POST"])
def submit():
    if "user_account" not in session:
        return redirect(url_for("login"))

    session["logged_in"] = True  # 🔹保證側邊欄顯示
    
    account = session["user_account"]
    name = session["user_name"]

    # 🔽🔽🔽 在這裡加入：更新「今天作答次數」 🔽🔽🔽
    today = date.today().isoformat()

    # 如果是新的一天，就重置
    if session.get("last_quiz_date") != today:
        session["last_quiz_date"] = today
        session["quiz_times_today"] = 0

    # 交卷算一次作答
    session["quiz_times_today"] = session.get("quiz_times_today", 0) + 1
    # 🔼🔼🔼 新增區塊到這裡 🔼🔼🔼
    
    score = 0
    details = []

    # 只批改表單裡有出現的題目 id
    for q in QUESTION_BANK:
        qid = q["id"]
        if qid in request.form:
            user_answer = request.form.get(qid)
            correct_answer = q["answer"]
            is_correct = (user_answer == correct_answer)
            if is_correct:
                score += 1

            details.append({
                "id": qid,
                "text": q["text"],
                "user_answer": user_answer if user_answer else "（未作答）",
                "correct_answer": correct_answer,
                "correct": is_correct,
                "explanation": q["explanation"]
            })

    total_questions = len(details)

        # ===== 寫入 quiz_results.xlsx（每列 = 一人一次作答） =====
    wb_r = load_workbook(RESULT_FILE)
    ws_r = wb_r["Results"]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 先計算這個學生是第幾次作答
    attempt_count = 0
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        acc_in_row = row[1]  # 第2欄是「帳號」
        if acc_in_row == account:
            attempt_count += 1
    attempt_no = attempt_count + 1  # 這次是第幾次作答

    # 先把這次作答的結果整理成 dict，方便填到對應欄位
    answer_map = {}  # key: 題目ID -> (答案字串, 是否正確)
    for d in details:
        answer_map[d["id"]] = (d["user_answer"], "O" if d["correct"] else "X")

    # 組一整列資料
    row_values = [
        now_str,       # 時間
        account,       # 帳號
        name,          # 姓名
        attempt_no,    # 作答次數
        score          # 本次分數
    ]

    # 依照 QUESTION_BANK 的順序，把每題填進去
    for q in QUESTION_BANK:
        qid = q["id"]
        if qid in answer_map:
            ans, mark = answer_map[qid]
        else:
            ans, mark = "", ""  # 這次沒出到的題目留空
        row_values.append(ans)
        row_values.append(mark)

    ws_r.append(row_values)
    wb_r.save(RESULT_FILE)
    

    # ===== 同步一份到 Google 試算表 =====
    try:
        sheet = get_google_sheet()
        sheet.append_row(row_values)
    except Exception as e:
        # 不讓學生看到錯誤，只在伺服器印出來方便你除錯
        print("寫入 Google Sheet 失敗：", e)


    # 更新使用者總積分
    wb_u, ws_u, row_idx = get_user_row(account)
    new_total_points = None
    if row_idx is not None:
        total_cell = ws_u.cell(row=row_idx, column=4)  # 第4欄是「總積分」
        current_total = total_cell.value or 0
        total_cell.value = current_total + score
        new_total_points = total_cell.value
        wb_u.save(USERS_FILE)
    else:
        # 理論上不會發生，如果 users.xlsx 沒這個人
        new_total_points = score

    session["total_points"] = new_total_points

    # 計算該生排名
    rank, total_users = get_user_rank(account)

    # 計算該生等級
    level = get_level(new_total_points)

    return render_template(
        "result.html",
        name=name,
        score=score,
        total=total_questions,
        details=details,
        total_points=new_total_points,
        rank=rank,
        total_users=total_users,
        level=level,
        show_explanation=SETTINGS.get("show_explanation", True)
    )


def _build_qid_meta():
    """把題庫轉成 {qid: {text, answer, explanation}} 方便查表。"""
    return {q["id"]: {"text": q["text"], "answer": q["answer"], "explanation": q.get("explanation", "")}
            for q in QUESTION_BANK}

@app.route("/review")
def review():
    if "user_account" not in session:
        return redirect(url_for("login"))

    account = session["user_account"]
    name = session.get("user_name", account)

    # 開啟成績檔
    wb_r = load_workbook(RESULT_FILE)
    ws_r = wb_r["Results"]

    # 讀標題列，找出各題「答案欄」與「對錯欄」的索引
    headers = [cell.value for cell in ws_r[1]]
    # 結構：時間, 帳號, 姓名, 作答次數, 本次分數, q1_答案, q1_是否正確, q2_答案, q2_是否正確, ...
    q_cols = {}  # qid -> (ans_idx, mark_idx)
    for i in range(5, len(headers), 2):  # 從第6欄(索引5)開始，每兩欄一題
        if i + 1 < len(headers):
            ans_h = headers[i]
            mark_h = headers[i + 1]
            # 期望欄名像 "q1_答案", "q1_是否正確"
            if ans_h and "_答案" in ans_h:
                qid = ans_h.split("_答案")[0]
                q_cols[qid] = (i, i + 1)

    qmeta = _build_qid_meta()

    # 蒐集「該生所有作答中答錯的題目」：統計錯題次數 & 最近一次錯誤
    wrong_map = {}  # qid -> {count, last_time, last_user_answer}
    for row in ws_r.iter_rows(min_row=2, values_only=True):
        if row[1] != account:
            continue
        # 時間字串
        tstr = row[0]
        try:
            tval = datetime.strptime(tstr, "%Y-%m-%d %H:%M:%S")
        except Exception:
            tval = None

        for qid, (ai, mi) in q_cols.items():
            mark = row[mi]  # "O" 或 "X" 或空
            if mark == "X":
                user_ans = row[ai]
                info = wrong_map.get(qid, {"count": 0, "last_time": None, "last_user_answer": ""})
                info["count"] += 1
                # 更新最近一次錯誤
                if tval and (info["last_time"] is None or tval > info["last_time"]):
                    info["last_time"] = tval
                    info["last_user_answer"] = user_ans
                wrong_map[qid] = info

    # 組成模板要用的清單
    wrong_list = []
    for qid, info in wrong_map.items():
        mm = qmeta.get(qid, {"text": f"{qid}（題庫已移除或未載入）", "answer": "", "explanation": ""})
        wrong_list.append({
            "id": qid,
            "text": mm["text"],
            "correct_answer": mm["answer"],
            "explanation": mm["explanation"],
            "wrong_count": info["count"],
            "last_time": info["last_time"].strftime("%Y-%m-%d %H:%M:%S") if info["last_time"] else "",
            "last_user_answer": info["last_user_answer"],
        })

    # 依最近錯誤時間(新到舊)排序
    wrong_list.sort(key=lambda x: x["last_time"], reverse=True)

    return render_template("review.html", name=name, wrong_list=wrong_list, title="錯題回顧")



import os
RUNNING_IN_RENDER = os.environ.get("RENDER") is not None

if __name__ == "__main__":
    if not RUNNING_IN_RENDER:  # 本機才會初始化
        init_users_excel()
        init_results_excel()

    app.run(host="0.0.0.0", port=5000, debug=True)

